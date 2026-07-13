import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_to_excel():
    with open('aqs_test_results.json') as f:
        data = json.load(f)
        
    records = []
    for x in data:
        name = x.get('name', '')
        existing = x.get('existing_aqs', 0.0)
        computed = x.get('computed_aqs', 0)
        diff = x.get('difference', 0.0)
        links = x.get('links', {})
        li = links.get('linkedin', '')
        gh = links.get('github', '')
        
        dims = x.get('dimensions', {})
        d1 = dims.get('D1', {}).get('points', 0)
        d2 = dims.get('D2', {}).get('points', 0)
        d3 = dims.get('D3', {}).get('points', 0)
        d4 = dims.get('D4', {}).get('points', 0)
        d5 = dims.get('D5', {}).get('points', 0)
        d6 = dims.get('D6', {}).get('points', 0)
        d7 = dims.get('D7', {}).get('points', 0)
        d8 = dims.get('D8', {}).get('points', 0)
        d9 = dims.get('D9', {}).get('points', 0)
        d10 = dims.get('D10', {}).get('points', 0)
        d11 = dims.get('D11', {}).get('points', 0)
        d12 = dims.get('D12', {}).get('points', 0)
        
        d1_just = dims.get('D1', {}).get('justification', '')
        d9_just = dims.get('D9', {}).get('justification', '')
        
        # Check if bypassed
        is_bypassed = "Bypassed LLM call" in d1_just
        
        records.append({
            "Author Name": name,
            "Original AQS": existing if existing > 0 else "N/A",
            "Computed AQS": computed,
            "Change (Computed - Original)": diff if existing > 0 else "N/A",
            "Score Source": "Bypassed (RL CoP File)" if is_bypassed else "Computed Fresh (Strict AQS V2)",
            "LinkedIn URL": li,
            "GitHub URL": gh,
            "D1 (Thought Leadership)": d1,
            "D2 (Writing Quality)": d2,
            "D3 (Tech Depth)": d3,
            "D4 (Topic Relevancy)": d4,
            "D5 (Industry Expert)": d5,
            "D6 (Industry Recognised)": d6,
            "D7 (Packt Affinity)": d7,
            "D8 (Adoption/Sales)": d8,
            "D9 (Reach)": d9,
            "D10 (Content Output)": d10,
            "D11 (Responsiveness)": d11,
            "D12 (Commerciality)": d12,
            "D1 Justification": d1_just,
            "Reach Justification": d9_just
        })
        
    df = pd.DataFrame(records)
    
    # Save as Excel
    excel_path = 'aqs_final_scores.xlsx'
    df.to_excel(excel_path, index=False, sheet_name="AQS Summary")
    
    # Styling using openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["AQS Summary"]
    
    # Format Headers
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
        
    wb.save(excel_path)
    print("Exported final styled spreadsheet to:", excel_path)

if __name__ == "__main__":
    export_to_excel()
